import time
import requests
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from decouple import config
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException
from bs4 import BeautifulSoup

EMAIL = config('EMAIL_ADDRESS')
URL = config('IMDB_URL')

chrome_driver_path = 'C:\Development\chromedriver'
driver = webdriver.Chrome(executable_path=chrome_driver_path)


def get_movie_links():
    driver.get(URL)
    time.sleep(3)
    all_links = []
    while True:
        title_links = driver.find_elements_by_css_selector(
            '.lister-item-header a')
        title_numbers = driver.find_elements_by_css_selector(
            '.lister-item-index')
        links = [link.get_property('href') for link in title_links]
        all_links.extend(links)
        time.sleep(5)
        try:
            driver.find_element_by_xpath(
                '//*[@id="ratings-container"]/div[102]/div/div/a[2]').click()
        except NoSuchElementException:
            print('No more pages')
            break
        time.sleep(5)
    driver.close()
    return all_links


movie_links = get_movie_links()
actors_dict = {}
directors_dict = {}

for link in movie_links:
    response = requests.get(link)
    soup = BeautifulSoup(response.text, 'html.parser')
    director_exists = soup.select_one(
        '.summary_text+ .credit_summary_item .inline')

    try:
        movie_data = soup.select_one('.originalTitle')
        movie_name = movie_data.text.replace(
            u'\xa0', u' ').replace('(original title)', '').strip()
        movie_year_data = soup.select_one('#titleYear')
        movie_year = movie_year_data.text
        movie = movie_name + movie_year
    except AttributeError:
        movie_data = soup.select_one('h1')
        movie = movie_data.text.replace(u'\xa0', u' ').strip()

    if director_exists.text == 'Director:':
        director_data = soup.select_one(
            '.summary_text+ .credit_summary_item a')
        director_name = director_data.text
        if director_name in directors_dict:
            directors_dict[director_name]['movies'] += 1
            directors_dict[director_name]['movies_list'].append(movie)
        else:
            directors_dict[director_name] = {
                'movies': 1, 'movies_list': [movie]}

    print(f"Getting data from {movie_links.index(link) + 1}. {movie}")
    actors_data = soup.select('.primary_photo+ td')
    actors = [actor.text.replace('\n', '').strip() for actor in actors_data]
    for actor_instance in actors:
        if actor_instance in actors_dict:
            actors_dict[actor_instance]['movies'] += 1
            actors_dict[actor_instance]['movies_list'].append(movie)
        else:
            actors_dict[actor_instance] = {'movies': 1, 'movies_list': [movie]}


def get_year(movie):
    return movie[-5:-1:]


for actor in actors_dict:
    actors_dict[actor]['movies_list'].sort(key=get_year)

for director in directors_dict:
    directors_dict[director]['movies_list'].sort(key=get_year)


wb = Workbook()
dest_filename = 'actors.xlsx'
ws = wb.active
ws.title = 'Actors List'
ws.append(['Actor', 'Movies Number', 'Movies List'])
for (actor, values) in actors_dict.items():
    ws.append([actor, values['movies'], str(values['movies_list']
                                            ).replace('[', '').replace(']', '').replace("'", '')])


ws2 = wb.create_sheet(title='Directors')
ws2.append(['Director', 'Movies', 'Movies List'])
for (director, values) in directors_dict.items():
    ws2.append([director, values['movies'], str(values['movies_list']
                                                ).replace('[', '').replace(']', '').replace("'", '')])

wb.save(filename=dest_filename)
